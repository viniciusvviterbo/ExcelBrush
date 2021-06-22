from PIL import Image
from progress.bar import ChargingBar
import argparse
import openpyxl
import os

# Configuracao dos argumentos
parser = argparse.ArgumentParser(description = 'Software to draw images on spreadsheets.')
parser.add_argument('-i', action = 'store', dest = 'image_file_path', default = '', required = True, help = 'Image to be drawned on a spreadsheet.')
parser.add_argument('-height', action = 'store', dest = 'image_height', default = '', required = False, help = 'Height in pixels of the spreadsheet image. Maximum of 1048576 pixels.')
parser.add_argument('-width', action = 'store', dest = 'image_width', default = '', required = False, help = 'Width in pixels of the spreadsheet image. Maximum of 1024 pixels.')

# Retorna a string hexadecimal correspondente a cor RGB informada
def rgb2hex(cor):
    return "{:02x}{:02x}{:02x}".format(cor[0], cor[1], cor[2])

def main():

    # Recebe os argumentos. Se as variaveis nao forem passadas, retorna -h
    arguments = parser.parse_args()
    imagem_original = Image.open(arguments.image_file_path)
    tmp_larg, tmp_alt = imagem_original.size # Retorna as dimensões da imagem informada
    imagem_altura = tmp_alt if (arguments.image_height == '') else int(arguments.image_height)
    imagem_largura = tmp_larg if (arguments.image_width == '') else int(arguments.image_width)   # assim como a maior largura possível para uma imagem será 1024 pixels

#    if(imagem_largura > 1024 || imagem_altura > 1048576):
#        raise Exception('Image is too big to be drawned. Choose a smaller resolution.')

    try:
        # Define o nome base dos arquivos a serem criados
        nome_arquivo = os.path.splitext(os.path.basename(arguments.image_file_path))[0] + str(imagem_altura) + 'x' + str(imagem_largura)
        # Cria uma imagem reduzida que é nada mais que a imagem_original redimensionada com a resolução informada
        imagem_reduzida = imagem_original.resize((imagem_largura, imagem_altura), Image.BILINEAR)
        # Os dados da imagem sao lidos e armazenados em 'pix' 
        pix = imagem_reduzida.load()
        # Armazena as dimensoes da imagem_reduzida
        largura = imagem_reduzida.size[0]
        altura = imagem_reduzida.size[1]
        
        # Informa o início do processo
        print('Processo iniciado.')

        # Instancia objetos para manipulacao da planilha
        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        # Inicia a interface da CLI que indica o progresso do processo
        barra_progresso = ChargingBar('Desenhando', max = largura)
        for x in range(1, largura + 1):
            # Altera a altura da linha
            # PARA ALTERAÇOES DO TAMANHO, MODIFICAR APENAS O VALOR QUE MULTIPLICA 18.00 PARA MANTER AS CELULAS QUADRADAS
            ws.row_dimensions[x].height = 18.00 * 1
        
            for y in range(1, altura + 1):
                # Altera a largura da coluna
                # PARA ALTERAÇOES DO TAMANHO, MODIFICAR APENAS O VALOR QUE MULTIPLICA 2.43 PARA MANTER AS CELULAS QUADRADAS
                ws.column_dimensions[ws.cell(x, y).column_letter].width = 2.43 * 1
            
                # Identifica a cor do pixel da iteracao
                cor = rgb2hex(pix[x - 1, y - 1]).upper()
                
                # Pinta a celula da planilha
                ws.cell(y, x).fill = openpyxl.styles.PatternFill(fgColor=cor, fill_type='solid')
            # Atualiza a barra de progresso
            barra_progresso.next()
        # Finaliza  barra de progresso
        barra_progresso.finish()

        # Define o nome do diretório destino das imagens
        diretorio_planilhas = 'Done_Files'
        
        # Cria o diretorio de arquivos prontos caso não exista
        if not os.path.exists(diretorio_planilhas):
            os.mkdir(diretorio_planilhas)

        # Salva a pixel art
        wb.save('./{}/'.format(diretorio_planilhas) + nome_arquivo + '.xlsx')

        # Informa o fim da execução
        print('Execução concluída!')

    except():
        print('An error occurred.')

# Chama a função main
if __name__ == '__main__':
    main()
